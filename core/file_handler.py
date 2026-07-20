# ──────────────────────────────────────────────────────────────────────────────
# file_handler.py - File Upload & Processing Utilities
# Secure file uploads, validation, storage, and processing
# ──────────────────────────────────────────────────────────────────────────────

import os
import logging
import mimetypes
from typing import Optional, BinaryIO, Tuple
from pathlib import Path
from datetime import datetime
import hashlib

logger = logging.getLogger(__name__)


class FileUploadConfig:
    """File upload configuration."""
    
    # File size limits (in bytes)
    MAX_UPLOAD_SIZE = 50 * 1024 * 1024  # 50 MB
    MAX_IMAGE_SIZE = 10 * 1024 * 1024   # 10 MB
    MAX_DOCUMENT_SIZE = 25 * 1024 * 1024  # 25 MB
    
    # Allowed file types
    ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/gif", "image/webp"}
    ALLOWED_DOCUMENT_TYPES = {
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "text/plain",
        "text/csv",
    }
    ALLOWED_VIDEO_TYPES = {"video/mp4", "video/quicktime"}
    
    # Upload directories
    UPLOADS_DIR = os.getenv("UPLOADS_DIR", "storage/uploads")
    TEMP_DIR = os.getenv("TEMP_DIR", "storage/temp")
    
    def __init__(self):
        # Create directories if they don't exist
        Path(self.UPLOADS_DIR).mkdir(parents=True, exist_ok=True)
        Path(self.TEMP_DIR).mkdir(parents=True, exist_ok=True)


class FileValidator:
    """Validate files before processing."""
    
    config = FileUploadConfig()
    
    @staticmethod
    def validate_file(
        filename: str,
        file_size: int,
        file_type: str,
    ) -> Tuple[bool, Optional[str]]:
        """Validate file upload."""
        # Check file size
        if file_size > FileUploadConfig.MAX_UPLOAD_SIZE:
            return False, f"File too large. Max: {FileUploadConfig.MAX_UPLOAD_SIZE} bytes"
        
        # Check file extension
        allowed_extensions = {".jpg", ".jpeg", ".png", ".gif", ".pdf", ".doc", ".docx", ".txt", ".csv"}
        file_ext = Path(filename).suffix.lower()
        
        if file_ext not in allowed_extensions:
            return False, f"File type not allowed: {file_ext}"
        
        # Check MIME type
        if file_type not in (
            FileUploadConfig.ALLOWED_IMAGE_TYPES |
            FileUploadConfig.ALLOWED_DOCUMENT_TYPES |
            FileUploadConfig.ALLOWED_VIDEO_TYPES
        ):
            return False, f"MIME type not allowed: {file_type}"
        
        # Check for suspicious filenames
        if any(c in filename for c in ["//", "\\", "..", "\x00"]):
            return False, "Invalid filename"
        
        return True, None
    
    @staticmethod
    def validate_image(filename: str, file_size: int, file_type: str) -> Tuple[bool, Optional[str]]:
        """Validate image file."""
        if file_size > FileUploadConfig.MAX_IMAGE_SIZE:
            return False, f"Image too large. Max: {FileUploadConfig.MAX_IMAGE_SIZE} bytes"
        
        if file_type not in FileUploadConfig.ALLOWED_IMAGE_TYPES:
            return False, f"Image type not allowed: {file_type}"
        
        return True, None
    
    @staticmethod
    def validate_document(filename: str, file_size: int, file_type: str) -> Tuple[bool, Optional[str]]:
        """Validate document file."""
        if file_size > FileUploadConfig.MAX_DOCUMENT_SIZE:
            return False, f"Document too large. Max: {FileUploadConfig.MAX_DOCUMENT_SIZE} bytes"
        
        if file_type not in FileUploadConfig.ALLOWED_DOCUMENT_TYPES:
            return False, f"Document type not allowed: {file_type}"
        
        return True, None

    @staticmethod
    def validate_file_content(content: bytes, filename: str) -> Tuple[bool, Optional[str]]:
        """Validate file content by checking magic bytes to prevent execution of malicious uploads."""
        if not content:
            return False, "File is empty"
            
        ext = Path(filename).suffix.lower()
        if ext == ".pdf":
            if not content.startswith(b"%PDF"):
                return False, "Invalid PDF: Magic bytes signature mismatch."
        elif ext == ".docx":
            if not content.startswith(b"PK\x03\x04"):
                return False, "Invalid DOCX: Magic bytes signature mismatch."
        elif ext == ".doc":
            # doc files can be OLE compound binary or sometimes standard zip format (if renamed docx)
            if not (content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1") or content.startswith(b"PK\x03\x04")):
                return False, "Invalid DOC: Magic bytes signature mismatch."
        elif ext == ".rtf":
            if not content.startswith(b"{\\rtf"):
                return False, "Invalid RTF: Magic bytes signature mismatch."
        elif ext == ".txt":
            # Simple heuristic check to confirm it's not a binary disguised as text
            null_count = content.count(b"\x00")
            if len(content) > 0 and (null_count / len(content)) > 0.05:
                return False, "Invalid TXT: High density of null bytes suggests binary payload."
                
        return True, None


class FileStorage:
    """Handle file storage operations."""
    
    config = FileUploadConfig()
    
    @staticmethod
    def generate_filename(original_filename: str) -> str:
        """Generate unique filename."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        file_hash = hashlib.md5(original_filename.encode()).hexdigest()[:8]
        ext = Path(original_filename).suffix.lower()
        return f"{timestamp}_{file_hash}{ext}"
    
    @staticmethod
    def save_file(
        file_content: BinaryIO,
        original_filename: str,
        subfolder: str = "general",
    ) -> Tuple[bool, Optional[str], Optional[str]]:
        """Save file to storage."""
        try:
            # Generate unique filename
            filename = FileStorage.generate_filename(original_filename)
            
            # Create subfolder if needed
            folder = Path(FileStorage.config.UPLOADS_DIR) / subfolder
            folder.mkdir(parents=True, exist_ok=True)
            
            # Save file
            file_path = folder / filename
            
            with open(file_path, "wb") as f:
                f.write(file_content.read())
            
            # Return relative path
            relative_path = f"{subfolder}/{filename}"
            logger.info(f"File saved: {relative_path}")
            
            return True, relative_path, None
        
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            return False, None, str(e)
    
    @staticmethod
    def get_file(file_path: str) -> Optional[BinaryIO]:
        """Retrieve file from storage."""
        try:
            full_path = Path(FileStorage.config.UPLOADS_DIR) / file_path
            
            # Security: Prevent path traversal
            if not str(full_path).startswith(str(Path(FileStorage.config.UPLOADS_DIR))):
                logger.warning(f"Path traversal attempt: {file_path}")
                return None
            
            if full_path.exists():
                return open(full_path, "rb")
            
            return None
        
        except Exception as e:
            logger.error(f"Failed to retrieve file: {e}")
            return None
    
    @staticmethod
    def delete_file(file_path: str) -> bool:
        """Delete file from storage."""
        try:
            full_path = Path(FileStorage.config.UPLOADS_DIR) / file_path
            
            # Security: Prevent path traversal
            if not str(full_path).startswith(str(Path(FileStorage.config.UPLOADS_DIR))):
                logger.warning(f"Path traversal attempt: {file_path}")
                return False
            
            if full_path.exists():
                full_path.unlink()
                logger.info(f"File deleted: {file_path}")
                return True
            
            return False
        
        except Exception as e:
            logger.error(f"Failed to delete file: {e}")
            return False
    
    @staticmethod
    def get_file_size(file_path: str) -> Optional[int]:
        """Get file size in bytes."""
        try:
            full_path = Path(FileStorage.config.UPLOADS_DIR) / file_path
            
            if full_path.exists():
                return full_path.stat().st_size
            
            return None
        
        except Exception as e:
            logger.error(f"Failed to get file size: {e}")
            return None


class FileProcessor:
    """Process uploaded files."""
    
    @staticmethod
    async def process_pdf(file_path: str) -> Optional[dict]:
        """Extract text from PDF."""
        try:
            import pdfplumber
            
            full_path = Path(FileUploadConfig.UPLOADS_DIR) / file_path
            
            if not full_path.exists():
                return None
            
            with pdfplumber.open(full_path) as pdf:
                text = ""
                for page in pdf.pages:
                    text += page.extract_text() or ""
            
            return {
                "text": text,
                "pages": len(pdf.pages),
            }
        
        except Exception as e:
            logger.error(f"Failed to process PDF: {e}")
            return None
    
    @staticmethod
    async def process_csv(file_path: str) -> Optional[list]:
        """Parse CSV file."""
        try:
            import csv
            
            full_path = Path(FileUploadConfig.UPLOADS_DIR) / file_path
            
            if not full_path.exists():
                return None
            
            rows = []
            with open(full_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)
            
            return rows
        
        except Exception as e:
            logger.error(f"Failed to process CSV: {e}")
            return None
    
    @staticmethod
    async def process_excel(file_path: str) -> Optional[dict]:
        """Parse Excel file."""
        try:
            import openpyxl
            
            full_path = Path(FileUploadConfig.UPLOADS_DIR) / file_path
            
            if not full_path.exists():
                return None
            
            workbook = openpyxl.load_workbook(full_path)
            sheets = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(list(row))
                sheets[sheet_name] = rows
            
            return sheets
        
        except Exception as e:
            logger.error(f"Failed to process Excel: {e}")
            return None


# Usage in FastAPI:
#
# from fastapi import UploadFile
#
# @app.post("/api/uploads/resume")
# async def upload_resume(file: UploadFile):
#     # Validate file
#     is_valid, error = FileValidator.validate_document(
#         file.filename,
#         len(await file.read()),
#         file.content_type,
#     )
#     
#     if not is_valid:
#         return error_response(message=error, status_code=400)
#     
#     # Save file
#     await file.seek(0)  # Reset pointer
#     success, file_path, error = FileStorage.save_file(
#         file.file,
#         file.filename,
#         subfolder="resumes"
#     )
#     
#     if not success:
#         return error_response(message=error, status_code=500)
#     
#     return success_response(data={"file_path": file_path})
