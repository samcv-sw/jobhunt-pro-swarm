# ──────────────────────────────────────────────────────────────────────────────
# export_utils.py - Data Export & Reporting Utilities
# Generate reports, export data in multiple formats (CSV, Excel, PDF)
# ──────────────────────────────────────────────────────────────────────────────

import csv
import json
import logging
from io import StringIO, BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


class ExportFormat:
    """Export format options."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    PDF = "pdf"


class ExportHelper:
    """Helper for exporting data in various formats."""
    
    @staticmethod
    def export_to_csv(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
    ) -> str:
        """Export data to CSV."""
        try:
            if not data:
                return ""
            
            output = StringIO()
            fieldnames = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            
            writer.writeheader()
            writer.writerows(data)
            
            csv_content = output.getvalue()
            
            # Save to file if filename provided
            if filename:
                Path("exports").mkdir(exist_ok=True)
                with open(f"exports/{filename}", "w", newline="", encoding="utf-8") as f:
                    f.write(csv_content)
                logger.info(f"CSV exported to {filename}")
            
            return csv_content
        
        except Exception as e:
            logger.error(f"Failed to export CSV: {e}")
            return ""
    
    @staticmethod
    def export_to_json(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        pretty: bool = True,
    ) -> str:
        """Export data to JSON."""
        try:
            json_content = json.dumps(data, indent=2 if pretty else None, default=str)
            
            # Save to file if filename provided
            if filename:
                Path("exports").mkdir(exist_ok=True)
                with open(f"exports/{filename}", "w", encoding="utf-8") as f:
                    f.write(json_content)
                logger.info(f"JSON exported to {filename}")
            
            return json_content
        
        except Exception as e:
            logger.error(f"Failed to export JSON: {e}")
            return ""
    
    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
    ) -> BytesIO:
        """Export data to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            
            if not data:
                return BytesIO()
            
            # Write header
            fieldnames = list(data[0].keys())
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, fieldname in enumerate(fieldnames, 1):
                cell = sheet.cell(row=1, column=col_idx, value=fieldname)
                cell.fill = header_fill
                cell.font = header_font
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, fieldname in enumerate(fieldnames, 1):
                    value = row_data.get(fieldname, "")
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col_idx, fieldname in enumerate(fieldnames, 1):
                max_length = len(str(fieldname))
                for row_data in data:
                    cell_value = str(row_data.get(fieldname, ""))
                    max_length = max(max_length, len(cell_value))
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Save to file if filename provided
            if filename:
                Path("exports").mkdir(exist_ok=True)
                with open(f"exports/{filename}", "wb") as f:
                    f.write(output.getvalue())
                logger.info(f"Excel exported to {filename}")
            
            return output
        
        except Exception as e:
            logger.error(f"Failed to export Excel: {e}")
            return BytesIO()
    
    @staticmethod
    def export_to_pdf(
        data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        title: str = "Report",
    ) -> Optional[bytes]:
        """Export data to PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []
            
            # Add title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=24,
                textColor=colors.HexColor("#366092"),
                spaceAfter=30,
            )
            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 0.3 * inch))
            
            if not data:
                doc.build(elements)
                return output.getvalue()
            
            # Create table
            fieldnames = list(data[0].keys())
            table_data = [fieldnames]
            
            for row_data in data:
                row = [str(row_data.get(field, "")) for field in fieldnames]
                table_data.append(row)
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            pdf_content = output.getvalue()
            
            # Save to file if filename provided
            if filename:
                Path("exports").mkdir(exist_ok=True)
                with open(f"exports/{filename}", "wb") as f:
                    f.write(pdf_content)
                logger.info(f"PDF exported to {filename}")
            
            return pdf_content
        
        except Exception as e:
            logger.error(f"Failed to export PDF: {e}")
            return None


    @staticmethod
    def export_to_latex(profile_data: Dict[str, Any]) -> str:
        """Export candidate profile to clean LaTeX resume format."""
        name = profile_data.get("name", "Candidate Name")
        title = profile_data.get("title", "Professional Software Engineer")
        email = profile_data.get("email", "email@example.com")
        summary = profile_data.get("summary", "Experienced engineer specializing in scalable software systems.")
        skills = ", ".join(profile_data.get("skills", ["Python", "FastAPI", "Next.js", "Docker"]))

        latex_doc = f"""\\documentclass[11pt,a4paper]{{article}}
\\usepackage[utf8]{{utf8}}
\\usepackage[margin=0.75in]{{geometry}}
\\usepackage{{hyperref}}

\\begin{{document}}
\\begin{{center}}
    {{\\Huge \\bfseries {name}}}\\\\[4pt]
    {{\\Large {title}}}\\\\[4pt]
    \\href{{mailto:{email}}}{{{email}}}
\\end{{center}}

\\hrule
\\vspace{{10pt}}

\\section*{{Professional Summary}}
{summary}

\\section*{{Core Skills}}
{skills}

\\section*{{Experience}}
\\textbf{{Senior Engineer}} -- \\textit{{JobHunt Pro Sovereign Systems}}\\\\
Developed autonomous high-throughput multi-agent automation engines.

\\end{{document}}
"""
        return latex_doc

    @staticmethod
    def export_to_html_portfolio(profile_data: Dict[str, Any]) -> str:
        """Export candidate profile to responsive HTML portfolio format."""
        name = profile_data.get("name", "Candidate Name")
        title = profile_data.get("title", "Software Engineer")
        summary = profile_data.get("summary", "Building next-generation sovereign SaaS ecosystems.")
        skills = profile_data.get("skills", ["Python", "FastAPI", "React", "Docker"])
        skills_html = "".join([f'<span class="skill-tag">{s}</span>' for s in skills])

        html_portfolio = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{name} - Portfolio</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #0f172a; color: #f8fafc; margin: 0; padding: 2rem; }}
        .card {{ max-width: 800px; margin: 0 auto; background: #1e293b; border-radius: 12px; padding: 2.5rem; box-shadow: 0 10px 25px rgba(0,0,0,0.5); }}
        h1 {{ color: #38bdf8; margin-bottom: 0.5rem; }}
        .title {{ color: #94a3b8; font-size: 1.2rem; margin-bottom: 1.5rem; }}
        .skill-tag {{ display: inline-block; background: #0284c7; color: white; padding: 4px 12px; border-radius: 16px; margin: 4px; font-size: 0.9rem; }}
    </style>
</head>
<body>
    <div class="card">
        <h1>{name}</h1>
        <div class="title">{title}</div>
        <p>{summary}</p>
        <h3>Skills & Expertise</h3>
        <div>{skills_html}</div>
    </div>
</body>
</html>
"""
        return html_portfolio


class ReportBuilder:
    """Build reports from data."""
    
    def __init__(self, title: str):
        self.title = title
        self.sections = []
        self.metadata = {
            "generated_at": datetime.utcnow().isoformat(),
            "title": title,
        }
    
    def add_section(self, name: str, data: List[Dict[str, Any]]) -> None:
        """Add report section."""
        self.sections.append({
            "name": name,
            "data": data,
        })
    
    def add_metadata(self, key: str, value: Any) -> None:
        """Add metadata field."""
        self.metadata[key] = value
    
    def get_summary(self) -> Dict[str, Any]:
        """Get report summary."""
        return {
            "title": self.title,
            "generated_at": self.metadata["generated_at"],
            "sections": len(self.sections),
            "total_rows": sum(len(s["data"]) for s in self.sections),
        }
    
    def export(self, format: str = "json", filename: Optional[str] = None) -> Any:
        """Export report."""
        report_data = {
            "metadata": self.metadata,
            "sections": self.sections,
        }
        
        if format == ExportFormat.JSON:
            return ExportHelper.export_to_json(
                report_data.get("sections", []),
                filename,
            )
        elif format == ExportFormat.CSV:
            # For CSV, flatten all sections
            all_data = []
            for section in self.sections:
                all_data.extend(section["data"])
            return ExportHelper.export_to_csv(all_data, filename)
        elif format == ExportFormat.EXCEL:
            all_data = []
            for section in self.sections:
                all_data.extend(section["data"])
            return ExportHelper.export_to_excel(all_data, filename)
        elif format == ExportFormat.PDF:
            all_data = []
            for section in self.sections:
                all_data.extend(section["data"])
            return ExportHelper.export_to_pdf(
                all_data,
                filename,
                title=self.title,
            )


class AuditLog:
    """Audit logging for compliance."""
    
    def __init__(self):
        self._logs: List[Dict[str, Any]] = []
    
    def log_action(
        self,
        action: str,
        user_id: int,
        resource_type: str,
        resource_id: Any,
        changes: Optional[Dict] = None,
        status: str = "success",
    ) -> None:
        """Log an action."""
        log_entry = {
            "timestamp": datetime.utcnow().isoformat(),
            "action": action,
            "user_id": user_id,
            "resource_type": resource_type,
            "resource_id": resource_id,
            "changes": changes,
            "status": status,
        }
        
        self._logs.append(log_entry)
        logger.info(f"Audit: {action} on {resource_type} {resource_id} by user {user_id}")
    
    def get_logs(
        self,
        user_id: Optional[int] = None,
        resource_type: Optional[str] = None,
        limit: int = 1000,
    ) -> List[Dict[str, Any]]:
        """Get audit logs."""
        logs = self._logs
        
        if user_id:
            logs = [l for l in logs if l["user_id"] == user_id]
        
        if resource_type:
            logs = [l for l in logs if l["resource_type"] == resource_type]
        
        return logs[-limit:]
    
    def export_logs(self, format: str = "json") -> Any:
        """Export audit logs."""
        if format == ExportFormat.CSV:
            return ExportHelper.export_to_csv(self._logs)
        elif format == ExportFormat.EXCEL:
            return ExportHelper.export_to_excel(self._logs)
        elif format == ExportFormat.PDF:
            return ExportHelper.export_to_pdf(self._logs, title="Audit Log")
        else:
            return ExportHelper.export_to_json(self._logs)


# Global audit log instance
audit_log = AuditLog()


# Usage in FastAPI:
#
# @app.get("/api/export/users")
# async def export_users(format: str = "csv"):
#     users = db.get_all_users()
#     
#     if format == "csv":
#         content = ExportHelper.export_to_csv(users)
#         return Response(content=content, media_type="text/csv")
#     elif format == "excel":
#         content = ExportHelper.export_to_excel(users)
#         return Response(content=content.getvalue(), media_type="application/vnd.ms-excel")
#     else:
#         content = ExportHelper.export_to_json(users)
#         return Response(content=content, media_type="application/json")
